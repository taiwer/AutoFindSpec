import openpyxl
import re

import tkinter as tk
from tkinter import filedialog

# Open and load the file
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

# Print the file name
print("你选择的文件是：", file_path)

# Load the Excel file into a Workbook object
workbook = openpyxl.load_workbook(file_path)

# 储存结果的二维数组
result = []

# 遍历每个 sheet
for sheet_name in workbook.sheetnames:
    # 获取当前 sheet
    sheet = workbook[sheet_name]

    # 遍历每一行
    for row_num, row in enumerate(sheet.iter_rows()):
        # 判断第一列是否为‘Site’
        if row[0].value == 'Site':
            # 将 sheet 名和行号存入二维数组
            result.append([sheet_name],row_num+1)

# 打印结果
print(result)


    # 循环输出第一列的值
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            print(cell.value)

print(type(workbook.sheetnames))
print(workbook.sheetnames)
print(workbook.active)


all_sheetnames = wb.sheetnames

now_sheet = wb.active
print('当前活动的是：'+str(now_sheet))

cell = wb['LIS15']
print(cell)

cell_10 = cell['A:D']
print(type(cell_10))

global logiclist

print('按行获取值')

global ord
ord:int = 0
for n in wb.sheetnames:
    for i in wb[n].iter_rows(min_row=3, max_row=3, min_col=8, max_col=8):
        for j in i:
            print(j.value)



global pageNum
pageNum = 9
log = open("./book.txt", "r+")

for n in wb.sheetnames:
    for i in wb[n].iter_rows(min_row=pageNum, max_row=pageNum, min_col=1, max_col=50):
        reg = re.compile(r'"(.*?)"')

        result = re.findall(reg, str(wb[n]))
        print('\n\n', file=log)
        print(wb[n], file=log)
        print('/*', '-'*80, '*/', file=log)
        print('/*', '-' * 20, ' ' * 11, result[0], '' * 2, 'Start', ' ' * 11, '-' * 20, '*/', file=log)
        print('/*', '-' * 80, '*/', file=log)
        print('proc sql;\n', f'  create table output.{result[0]} as\n', '         select', file=log)
        for j in i:
            if j.value == 'Comment' or j.value is None:
                break
            var_name = ''.join(re.findall(r'[A-Za-z0-9]', str(j.value)))
            if var_name == 'SITENAME' or var_name == 'siteid':
                var_name = 'SITEID'
            if var_name == 'FOLDERNAME' or var_name == 'visit':
                var_name = 'FOLDER_NAME'
            if var_name == 'FORMNAME' or var_name == 'Form':
                var_name = 'FORM_NAME'
            print(f'                ,{var_name} \'{j.value}\'', file=log)  # write to list
        print('         from rawdata.\n         where\n         order by siteid, subjid, spid\n', ';quit;', file=log)
        print('/*', '-'*22, ' '*11, result[0], ''*2, 'End', ' '*11, '-'*22, '*/', file=log)
        # if ord <= len(logiclist):
log.close()

